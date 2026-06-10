import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

INPUT_FILE = r"C:\Tech\YFBLandingPortal\customers_raw.txt"
OUTPUT_FILE = r"C:\Tech\YFBLandingPortal\customers_deduped.xlsx"


def normalize(name: str) -> str:
    name = name.lower()
    name = re.sub(r"[.,]", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def main():
    seen = {}
    rows = []

    with open(INPUT_FILE, encoding="utf-8") as f:
        lines = f.readlines()

    # Skip header line
    data_lines = lines[1:] if lines else []
    input_count = len(data_lines)

    for line in data_lines:
        parts = line.rstrip("\n").split("\t")
        if len(parts) < 2:
            continue
        name, phone = parts[0].strip(), parts[1].strip()
        key = normalize(name)
        if key not in seen:
            seen[key] = True
            rows.append((name, phone))

    duplicates_removed = input_count - len(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    header_font = Font(name="Arial", bold=True)
    header_fill = PatternFill("solid", start_color="ADD8E6")  # light blue

    ws.append(["Name", "Phone"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for name, phone in rows:
        ws.append([name, phone])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    wb.save(OUTPUT_FILE)

    print(f"Input rows   : {input_count}")
    print(f"Duplicates   : {duplicates_removed}")
    print(f"Output rows  : {len(rows)}")
    print(f"Saved to     : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
